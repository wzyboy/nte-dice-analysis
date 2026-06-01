import re
from pathlib import Path
from zipfile import ZIP_DEFLATED
from zipfile import ZipFile
from zipfile import ZipInfo
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import Side
from openpyxl.styles import Border
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .models import Record
from .layouts import is_arc_pool_type
from .constants import A_CLASS
from .constants import XLSX_HEADERS
from .constants import ARC_XLSX_HEADERS
from .export_records import records_by_pool
from .export_records import is_s_class_target
from .export_records import total_pull_counts
from .export_records import split_item_type_name
from .export_records import pulls_since_last_s_target
from .export_records import pulls_since_last_s_character

S_CHARACTER_FILL = PatternFill(fill_type='solid', fgColor='FCE7A1')
A_CLASS_FILL = PatternFill(fill_type='solid', fgColor='E9D5FF')
OTHER_ROW_FILL = PatternFill(fill_type='solid', fgColor='E5E7EB')
HEADER_FILL = PatternFill(fill_type='solid', fgColor='1F2937')
HEADER_FONT = Font(color='FFFFFF', bold=True)
THIN_BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)
DETERMINISTIC_XLSX_DATETIME = datetime(2000, 1, 1, 0, 0, 0)
DETERMINISTIC_ZIP_TIMESTAMP = (2000, 1, 1, 0, 0, 0)


def write_xlsx(path: Path, records: list[Record]) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for pool_type, pool_records in records_by_pool(records).items():
        write_pool_sheet(workbook, pool_type or 'unknown pool', pool_records)

    if not workbook.sheetnames:
        write_pool_sheet(workbook, 'Records', [])

    workbook.properties.created = DETERMINISTIC_XLSX_DATETIME
    workbook.properties.modified = DETERMINISTIC_XLSX_DATETIME
    workbook.save(path)
    normalize_xlsx_archive(path)


def write_pool_sheet(
    workbook: Workbook,
    pool_type: str,
    records: list[Record],
) -> None:
    sheet = workbook.create_sheet(safe_sheet_title(pool_type, workbook.sheetnames))
    if is_arc_pool_type(pool_type):
        write_arc_pool_sheet(sheet, records)
        return

    write_dice_pool_sheet(sheet, records)


def write_dice_pool_sheet(sheet: Worksheet, records: list[Record]) -> None:
    sheet.append(XLSX_HEADERS)

    display_records = list(reversed(records))
    pulls_since_last_s = pulls_since_last_s_character(display_records)
    total_pulls = total_pull_counts(display_records)
    for record, pulls_since, total_pull_count in zip(
        display_records,
        pulls_since_last_s,
        total_pulls,
        strict=True,
    ):
        item_type, item_name = split_item_type_name(record.item_name)
        sheet.append(
            [
                record.roll_points,
                item_type,
                item_name,
                record.rarity,
                quantity_value(record.quantity),
                datetime_value(record.obtained_at),
                pulls_since,
                total_pull_count,
            ],
        )

    style_sheet(
        sheet,
        display_records,
        header_count=len(XLSX_HEADERS),
        date_column=6,
        count_columns=(7, 8),
        widths=[14, 12, 24, 12, 10, 22, 12, 12],
    )


def write_arc_pool_sheet(sheet: Worksheet, records: list[Record]) -> None:
    sheet.append(ARC_XLSX_HEADERS)

    display_records = list(reversed(records))
    pulls_since_last_s = pulls_since_last_s_target(display_records)
    total_pulls = total_pull_counts(display_records)
    for record, pulls_since, total_pull_count in zip(
        display_records,
        pulls_since_last_s,
        total_pulls,
        strict=True,
    ):
        sheet.append(
            [
                record.research_type,
                record.item_name,
                record.rarity,
                datetime_value(record.obtained_at),
                pulls_since,
                total_pull_count,
            ],
        )

    style_sheet(
        sheet,
        display_records,
        header_count=len(ARC_XLSX_HEADERS),
        date_column=4,
        count_columns=(5, 6),
        widths=[16, 24, 12, 22, 12, 12],
    )


def safe_sheet_title(title: str, existing_titles: list[str]) -> str:
    cleaned = ''.join('_' if char in r'[]:*?/\\' else char for char in title)[:31] or 'Sheet'
    if cleaned not in existing_titles:
        return cleaned

    suffix = 2
    while True:
        suffix_text = f' {suffix}'
        candidate = f'{cleaned[: 31 - len(suffix_text)]}{suffix_text}'
        if candidate not in existing_titles:
            return candidate
        suffix += 1


def fill_for_record(record: Record) -> PatternFill:
    if record.rarity == A_CLASS:
        return A_CLASS_FILL
    if is_s_class_target(record):
        return S_CHARACTER_FILL
    return OTHER_ROW_FILL


def quantity_value(value: str) -> int | str:
    try:
        return int(value)
    except ValueError:
        return value


def datetime_value(value: str) -> datetime | str:
    try:
        return datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
    except ValueError:
        return value


def style_sheet(
    sheet: Worksheet,
    records: list[Record],
    *,
    header_count: int,
    date_column: int,
    count_columns: tuple[int, ...],
    widths: list[int],
) -> None:
    sheet.freeze_panes = 'A2'
    sheet.sheet_view.showGridLines = False

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    for row_index, record in enumerate(records, start=2):
        fill = fill_for_record(record)
        for cell in sheet[row_index]:
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center')
        sheet.cell(row=row_index, column=date_column).number_format = 'yyyy-mm-dd hh:mm:ss'
        for column_index in count_columns:
            sheet.cell(row=row_index, column=column_index).number_format = '0'

    set_column_widths(sheet, widths)
    sheet.auto_filter.ref = f'A1:{get_column_letter(header_count)}{max(sheet.max_row, 1)}'


def normalize_xlsx_archive(path: Path) -> None:
    with ZipFile(path, 'r') as source:
        entries = [
            (info, normalize_xlsx_entry(info.filename, source.read(info.filename))) for info in source.infolist()
        ]

    temp_path = path.with_suffix(f'{path.suffix}.tmp')
    with ZipFile(temp_path, 'w', compression=ZIP_DEFLATED, compresslevel=6) as target:
        for source_info, data in entries:
            info = ZipInfo(source_info.filename, DETERMINISTIC_ZIP_TIMESTAMP)
            info.compress_type = ZIP_DEFLATED
            info.external_attr = source_info.external_attr
            target.writestr(info, data)

    temp_path.replace(path)


def normalize_xlsx_entry(filename: str, data: bytes) -> bytes:
    if filename != 'docProps/core.xml':
        return data

    timestamp = '2000-01-01T00:00:00Z'
    text = data.decode('utf-8')
    text = re.sub(
        r'(<dcterms:created\b[^>]*>).*?(</dcterms:created>)',
        rf'\g<1>{timestamp}\2',
        text,
    )
    text = re.sub(
        r'(<dcterms:modified\b[^>]*>).*?(</dcterms:modified>)',
        rf'\g<1>{timestamp}\2',
        text,
    )
    return text.encode('utf-8')


def set_column_widths(sheet: Worksheet, widths: list[int]) -> None:
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width
