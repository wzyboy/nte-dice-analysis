from pathlib import Path
from collections.abc import Callable

from PIL import Image
from openpyxl import load_workbook

from nte_dice_analysis.png import write_png
from nte_dice_analysis.png import history_color
from nte_dice_analysis.png import summarize_pool
from nte_dice_analysis.xlsx import write_xlsx
from nte_dice_analysis.xlsx import safe_sheet_title
from nte_dice_analysis.models import Record
from nte_dice_analysis.constants import GIFT_ROLL_POINTS
from nte_dice_analysis.export_records import records_by_pool
from nte_dice_analysis.export_records import total_pull_counts
from nte_dice_analysis.export_records import split_item_type_name
from nte_dice_analysis.export_records import pulls_since_last_s_character


def test_record_output_round_trip(record_factory: Callable[..., Record]) -> None:
    record = record_factory(source_image='debug/page.png', page_row=3, confidence=0.876)

    row = record.to_output_row()
    restored = Record.from_output_row(row)

    assert row['page_row'] == '3'
    assert row['confidence'] == '0.876'
    assert restored == Record(
        pool_type=record.pool_type,
        source_image=Path('debug/page.png'),
        page_row=3,
        roll_points=record.roll_points,
        item_name=record.item_name,
        rarity=record.rarity,
        item_name_raw=record.item_name_raw,
        quantity=record.quantity,
        obtained_at=record.obtained_at,
        obtained_at_raw=record.obtained_at_raw,
        confidence=0.876,
    )


def test_xlsx_grouping_and_sheet_helpers(record_factory: Callable[..., Record]) -> None:
    records = [
        record_factory(pool_type='限定棋盘'),
        record_factory(pool_type='标准棋盘'),
    ]

    assert list(records_by_pool(records)) == ['限定棋盘', '标准棋盘']
    assert split_item_type_name('角色·薄荷') == ('角色', '薄荷')
    assert split_item_type_name('未知') == ('', '未知')
    assert safe_sheet_title('bad[]:*?/\\name', []) == 'bad_______name'
    assert safe_sheet_title('限定棋盘', ['限定棋盘']) == '限定棋盘 2'


def test_pulls_since_last_s_character_counts_from_oldest_record(
    record_factory: Callable[..., Record],
) -> None:
    records = [
        record_factory(page_row=4, roll_points=GIFT_ROLL_POINTS, item_name='道具·赠礼', rarity='B-Class'),
        record_factory(page_row=3, roll_points='1', item_name='道具·质实骰子', rarity='B-Class'),
        record_factory(page_row=2, roll_points='2', item_name='角色·薄荷', rarity='S-Class'),
        record_factory(page_row=1, roll_points='3', item_name='角色·哈尼娅', rarity='A-Class'),
    ]

    assert pulls_since_last_s_character(records) == [None, 1, 2, 1]


def test_total_pull_counts_skip_gifts(record_factory: Callable[..., Record]) -> None:
    records = [
        record_factory(page_row=4, roll_points=GIFT_ROLL_POINTS, item_name='道具·赠礼'),
        record_factory(page_row=3, roll_points='1'),
        record_factory(page_row=2, roll_points='2'),
        record_factory(page_row=1, roll_points=GIFT_ROLL_POINTS, item_name='道具·赠礼'),
        record_factory(page_row=0, roll_points='3'),
    ]

    assert total_pull_counts(records) == [None, 1, 2, None, 3]


def test_summarize_pool_skips_gifts_and_tracks_current_pity(
    record_factory: Callable[..., Record],
) -> None:
    records = [
        record_factory(
            page_row=1,
            roll_points='2',
            item_name='角色·哈尼娅',
            rarity='A-Class',
            obtained_at='2026-01-04 03:04:05',
        ),
        record_factory(
            page_row=2,
            roll_points=GIFT_ROLL_POINTS,
            item_name='道具·赠礼',
            rarity='B-Class',
            obtained_at='2026-01-03 03:04:05',
        ),
        record_factory(
            page_row=3,
            roll_points='1',
            item_name='角色·薄荷',
            rarity='S-Class',
            obtained_at='2026-01-02 03:04:05',
        ),
    ]

    summary = summarize_pool('限定棋盘', records)

    assert summary.total_pulls == 2
    assert summary.date_start == '2026-01-02'
    assert summary.date_end == '2026-01-04'
    assert summary.current_pity == 1
    assert [stat.count for stat in summary.rarity_stats] == [1, 1, 0]
    assert [(item.name, item.pulls) for item in summary.s_history] == [('薄荷', 1)]
    assert summary.average_s_pulls == 1


def test_history_color_is_deterministic_and_varied() -> None:
    assert history_color('薄荷') == history_color('薄荷')
    assert history_color('薄荷') != history_color('哈尼娅')


def test_write_xlsx_creates_pool_sheet(
    tmp_path: Path,
    record_factory: Callable[..., Record],
) -> None:
    path = tmp_path / 'records.xlsx'
    records = [
        record_factory(
            page_row=1,
            roll_points='3',
            item_name='角色·哈尼娅',
            rarity='A-Class',
            obtained_at='2026-01-02 03:04:07',
        ),
        record_factory(
            page_row=2,
            roll_points=GIFT_ROLL_POINTS,
            item_name='道具·赠礼',
            rarity='B-Class',
            obtained_at='2026-01-02 03:04:06',
        ),
        record_factory(
            page_row=3,
            roll_points='1',
            item_name='角色·薄荷',
            rarity='S-Class',
            obtained_at='2026-01-02 03:04:05',
        ),
    ]

    write_xlsx(path, records)

    workbook = load_workbook(path)
    sheet = workbook['限定棋盘']
    assert sheet['A1'].value == '投掷点数'
    assert sheet['H1'].value == '总抽数'
    assert sheet['C2'].value == '薄荷'
    assert sheet['C3'].value == '赠礼'
    assert sheet['C4'].value == '哈尼娅'
    assert sheet['G2'].value == 1
    assert sheet['H2'].value == 1
    assert sheet['G3'].value is None
    assert sheet['H3'].value is None
    assert sheet['G4'].value == 1
    assert sheet['H4'].value == 2


def test_write_xlsx_applies_requested_row_fills(
    tmp_path: Path,
    record_factory: Callable[..., Record],
) -> None:
    path = tmp_path / 'records.xlsx'
    records = [
        record_factory(item_name='角色·哈尼娅', rarity='A-Class'),
        record_factory(roll_points=GIFT_ROLL_POINTS, item_name='道具·赠礼', rarity='B-Class'),
        record_factory(item_name='角色·薄荷', rarity='S-Class'),
        record_factory(item_name='道具·质实骰子', rarity='S-Class'),
    ]

    write_xlsx(path, records)

    workbook = load_workbook(path)
    sheet = workbook['限定棋盘']
    assert sheet['C2'].value == '质实骰子'
    assert sheet['A2'].fill.fgColor.rgb == '00E5E7EB'
    assert sheet['C3'].value == '薄荷'
    assert sheet['A3'].fill.fgColor.rgb == '00FCE7A1'
    assert sheet['C4'].value == '赠礼'
    assert sheet['A4'].fill.fgColor.rgb == '00E5E7EB'
    assert sheet['C5'].value == '哈尼娅'
    assert sheet['A5'].fill.fgColor.rgb == '00E9D5FF'


def test_write_png_creates_summary_image(
    tmp_path: Path,
    record_factory: Callable[..., Record],
) -> None:
    path = tmp_path / 'records.png'
    records = [
        record_factory(pool_type='限定棋盘', item_name='角色·薄荷', rarity='S-Class'),
        record_factory(pool_type='标准棋盘', item_name='角色·哈尼娅', rarity='A-Class'),
    ]

    write_png(path, records)

    with Image.open(path) as image:
        assert image.format == 'PNG'
        assert image.size[0] >= 800
        assert image.size[1] >= 700
