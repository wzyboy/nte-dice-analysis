import re
from pathlib import Path
from collections.abc import Callable

import pytest
from PIL import Image
from openpyxl import load_workbook

from nte_dice_analysis import crop_cli
from nte_dice_analysis import recognize_cli
from nte_dice_analysis import export_png_cli
from nte_dice_analysis import export_xlsx_cli
from nte_dice_analysis import check_known_items_cli
from nte_dice_analysis.io import load_json
from nte_dice_analysis.io import write_json
from nte_dice_analysis.models import Record
from nte_dice_analysis.models import OcrPrediction


class PoolTypeOcr:
    def __init__(self) -> None:
        self.image_sizes: list[tuple[int, int]] = []

    def predict(self, image: object) -> list[OcrPrediction]:
        size = getattr(image, 'shape')
        self.image_sizes.append((int(size[1]), int(size[0])))
        return [
            {
                'rec_texts': ['标准棋盘'],
                'rec_scores': [0.99],
            },
        ]


class TableOcr:
    def predict(self, image: object) -> list[OcrPrediction]:
        return [
            {
                'rec_texts': ['角色', '·薄荷', 'x1', '2026年5月7日03:04:05'],
                'rec_scores': [0.95, 0.90, 0.85, 0.80],
                'rec_boxes': [
                    (230, 10, 270, 20),
                    (280, 10, 330, 20),
                    (520, 10, 560, 20),
                    (700, 10, 900, 20),
                ],
            },
        ]


class MissingTimestampTableOcr:
    def predict(self, image: object) -> list[OcrPrediction]:
        return [
            {
                'rec_texts': ['角色', '·薄荷', 'x1'],
                'rec_scores': [0.95, 0.90, 0.85],
                'rec_boxes': [
                    (230, 10, 270, 20),
                    (280, 10, 330, 20),
                    (520, 10, 560, 20),
                ],
            },
        ]


def test_crop_cli_writes_named_table_crop(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / '2026-05-25_21-06-03_NTE.png'
    Image.new('RGB', (100, 80), 'white').save(source)
    fake_ocr = PoolTypeOcr()
    monkeypatch.setattr(crop_cli, 'create_ocr', lambda options: fake_ocr)

    crop_cli.main(
        [
            str(source),
            '--table-crop',
            '10,20,50,60',
            '--pool-crop',
            '70,10,90,30',
        ],
    )

    output = tmp_path / '2026-05-25_21-06-03_NTE.table.标准棋盘.png'
    assert output.exists()
    assert Image.open(output).size == (40, 40)
    assert fake_ocr.image_sizes == [(20, 20)]


def test_crop_cli_skips_existing_table_crop_without_ocr(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    source = tmp_path / '2026-05-25_21-06-03_NTE.png'
    output = tmp_path / '2026-05-25_21-06-03_NTE.table.标准棋盘.png'
    Image.new('RGB', (100, 80), 'white').save(source)
    Image.new('RGB', (40, 40), 'black').save(output)
    monkeypatch.setattr(crop_cli, 'create_ocr', lambda options: pytest.fail('OCR should not be initialized'))

    crop_cli.main([str(tmp_path)])

    assert Image.open(output).getpixel((0, 0)) == (0, 0, 0)
    assert 'wrote 0 cropped table images; skipped 1 existing files' in capsys.readouterr().out


def test_crop_cli_rejects_device_option(tmp_path: Path) -> None:
    source = tmp_path / 'source.png'
    Image.new('RGB', (100, 80), 'white').save(source)

    with pytest.raises(SystemExit) as error:
        crop_cli.parse_args([str(source), '--device', 'gpu:0'])

    assert error.value.code == 2


def test_recognize_cli_writes_json_for_cropped_table(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    table = tmp_path / '2026-05-25_21-06-03_NTE.table.标准棋盘.png'
    Image.new('RGB', (1000, 100), 'white').save(table)
    monkeypatch.setattr(recognize_cli, 'create_ocr', lambda options: TableOcr())

    recognize_cli.main(
        [
            str(table),
            '--row-count',
            '2',
            '--row-top',
            '0',
            '--row-bottom',
            '1',
        ],
    )

    records = load_json(table.with_suffix('.json'))
    assert len(records) == 1
    assert records[0].pool_type == '标准棋盘'
    assert records[0].source_image == table
    assert records[0].item_name == '角色·薄荷'
    assert records[0].obtained_at == '2026-05-07 03:04:05'


def test_recognize_cli_rejects_device_option(tmp_path: Path) -> None:
    table = tmp_path / 'table.png'
    Image.new('RGB', (1000, 100), 'white').save(table)

    with pytest.raises(SystemExit) as error:
        recognize_cli.parse_args([str(table), '--device', 'gpu:0'])

    assert error.value.code == 2


def test_recognize_cli_rejects_missing_timestamp(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    table = tmp_path / '2026-05-25_21-06-03_NTE.table.标准棋盘.png'
    json_out = table.with_suffix('.json')
    Image.new('RGB', (1000, 100), 'white').save(table)
    monkeypatch.setattr(recognize_cli, 'create_ocr', lambda options: MissingTimestampTableOcr())

    with pytest.raises(SystemExit, match='missing obtained_at'):
        recognize_cli.main(
            [
                str(table),
                '--row-count',
                '2',
                '--row-top',
                '0',
                '--row-bottom',
                '1',
            ],
        )

    assert not json_out.exists()


def test_recognize_cli_skips_existing_json_without_ocr(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
    record_factory: Callable[..., Record],
) -> None:
    table = tmp_path / '2026-05-25_21-06-03_NTE.table.标准棋盘.png'
    json_out = table.with_suffix('.json')
    Image.new('RGB', (1000, 100), 'white').save(table)
    write_json(json_out, [record_factory(source_image=str(table))])
    monkeypatch.setattr(recognize_cli, 'create_ocr', lambda options: pytest.fail('OCR should not be initialized'))

    recognize_cli.main([str(table)])

    assert len(load_json(json_out)) == 1
    assert 'wrote 0 records to 0 JSON files; skipped 1 existing files' in capsys.readouterr().out


def test_recognize_cli_requires_pool_type(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    table = tmp_path / 'cropped.png'
    Image.new('RGB', (1000, 100), 'white').save(table)
    monkeypatch.setattr(recognize_cli, 'create_ocr', lambda options: TableOcr())

    with pytest.raises(SystemExit, match='could not infer pool type'):
        recognize_cli.main([str(table)])


def test_export_xlsx_cli_deduplicates_json_inputs(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
    record_factory: Callable[..., Record],
) -> None:
    first_json = tmp_path / 'page1.json'
    second_json = tmp_path / 'page2.json'
    xlsx_out = tmp_path / 'records.xlsx'
    first = record_factory(source_image='page1.png', item_name='角色·娜娜莉', confidence=0.1)
    better = record_factory(source_image='page2.png', item_name='角色·娜娜莉', confidence=0.9)
    write_json(first_json, [first])
    write_json(second_json, [better])

    export_xlsx_cli.main([str(first_json), str(second_json), '--xlsx-out', str(xlsx_out)])

    workbook = load_workbook(xlsx_out)
    sheet = workbook['限定棋盘']
    assert sheet.max_row == 2
    assert sheet['C2'].value == '娜娜莉'
    assert f'loaded 2 records from 2 JSON files; wrote 1 records to {xlsx_out}' in capsys.readouterr().out


def test_export_xlsx_cli_rejects_missing_timestamp(
    tmp_path: Path,
    record_factory: Callable[..., Record],
) -> None:
    json_in = tmp_path / 'records.json'
    xlsx_out = tmp_path / 'records.xlsx'
    write_json(json_in, [record_factory(obtained_at='')])

    with pytest.raises(SystemExit, match='missing obtained_at'):
        export_xlsx_cli.main([str(json_in), '--xlsx-out', str(xlsx_out)])

    assert not xlsx_out.exists()


def test_export_xlsx_cli_rejects_invalid_pull_groups(
    tmp_path: Path,
    record_factory: Callable[..., Record],
) -> None:
    json_in = tmp_path / 'records.json'
    xlsx_out = tmp_path / 'records.xlsx'
    records = [record_factory(roll_points=str((index % 6) + 1), item_name=f'角色·{index}') for index in range(10)]
    write_json(json_in, records)

    with pytest.raises(SystemExit, match='invalid pull groups'):
        export_xlsx_cli.main([str(json_in), '--xlsx-out', str(xlsx_out)])

    assert not xlsx_out.exists()


def test_export_xlsx_cli_rejects_no_dedup_flag(
    tmp_path: Path,
    record_factory: Callable[..., Record],
) -> None:
    json_in = tmp_path / 'records.json'
    xlsx_out = tmp_path / 'records.xlsx'
    records = [record_factory(roll_points=str((index % 6) + 1), item_name=f'角色·{index}') for index in range(10)]
    write_json(json_in, records)

    with pytest.raises(SystemExit) as error:
        export_xlsx_cli.main([str(json_in), '--xlsx-out', str(xlsx_out), '--no-dedup'])

    assert error.value.code == 2
    assert not xlsx_out.exists()


def test_export_png_cli_deduplicates_json_inputs(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
    record_factory: Callable[..., Record],
) -> None:
    first_json = tmp_path / 'page1.json'
    second_json = tmp_path / 'page2.json'
    png_out = tmp_path / 'records.png'
    first = record_factory(source_image='page1.png', item_name='角色·娜娜莉', confidence=0.1)
    better = record_factory(source_image='page2.png', item_name='角色·娜娜莉', confidence=0.9)
    write_json(first_json, [first])
    write_json(second_json, [better])

    export_png_cli.main([str(first_json), str(second_json), '--png-out', str(png_out)])

    assert png_out.exists()
    with Image.open(png_out) as image:
        assert image.format == 'PNG'
    output = capsys.readouterr().out
    assert f'loaded 2 records from 2 JSON files; wrote 1 records to {png_out}' in output
    assert '一共 1 抽 已累计 0 抽未出 S-Class 角色' in output
    assert 'S-Class 角色历史记录: 娜娜莉[1]' in output
    assert 'S-Class 角色平均出货次数为: 1' in output


def test_export_png_cli_rejects_missing_timestamp(
    tmp_path: Path,
    record_factory: Callable[..., Record],
) -> None:
    json_in = tmp_path / 'records.json'
    png_out = tmp_path / 'records.png'
    write_json(json_in, [record_factory(obtained_at='')])

    with pytest.raises(SystemExit, match='missing obtained_at'):
        export_png_cli.main([str(json_in), '--png-out', str(png_out)])

    assert not png_out.exists()


def test_export_png_cli_rejects_no_dedup_flag(tmp_path: Path) -> None:
    json_in = tmp_path / 'records.json'
    png_out = tmp_path / 'records.png'
    json_in.write_text('[]', encoding='utf-8')

    with pytest.raises(SystemExit) as error:
        export_png_cli.main([str(json_in), '--png-out', str(png_out), '--no-dedup'])

    assert error.value.code == 2
    assert not png_out.exists()


def test_check_known_items_cli_accepts_known_items(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
    record_factory: Callable[..., Record],
) -> None:
    json_in = tmp_path / 'records.json'
    write_json(json_in, [record_factory(item_name='角色·娜娜莉')])

    check_known_items_cli.main([str(json_in)])

    assert 'all item names are present in known items; checked 1 records from 1 JSON files' in capsys.readouterr().out


def test_check_known_items_cli_reports_missing_items(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
    record_factory: Callable[..., Record],
) -> None:
    json_in = tmp_path / 'records.json'
    write_json(
        json_in,
        [
            record_factory(source_image='page1.png', page_row=1, item_name='角色·新角色'),
            record_factory(source_image='page2.png', page_row=2, item_name='角色·新角色'),
        ],
    )

    with pytest.raises(SystemExit) as error:
        check_known_items_cli.main([str(json_in)])

    assert error.value.code == 1
    output = capsys.readouterr().out
    assert 'missing known items:' in output
    assert '- 角色·新角色 (2 occurrences)' in output
    assert f'{json_in} (page1.png, row 1)' in output
    assert f'{json_in} (page2.png, row 2)' in output


def test_check_known_items_cli_resolves_directory_inputs(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
    record_factory: Callable[..., Record],
) -> None:
    write_json(tmp_path / 'first.json', [record_factory(item_name='角色·娜娜莉')])
    write_json(tmp_path / 'second.json', [record_factory(item_name='弧盘·「我们。」')])

    check_known_items_cli.main([str(tmp_path)])

    assert 'checked 2 records from 2 JSON files' in capsys.readouterr().out


def test_check_known_items_cli_uses_custom_known_items(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
    record_factory: Callable[..., Record],
) -> None:
    known_items = tmp_path / 'known_items.txt'
    known_items.write_text('角色·新角色\n', encoding='utf-8')
    json_in = tmp_path / 'records.json'
    write_json(json_in, [record_factory(item_name='角色·新角色')])

    check_known_items_cli.main([str(json_in), '--known-items', str(known_items)])

    assert 'checked 1 records from 1 JSON files' in capsys.readouterr().out


def test_check_known_items_cli_reports_missing_json(tmp_path: Path) -> None:
    missing_json = tmp_path / 'missing.json'

    with pytest.raises(SystemExit, match=re.escape(f'JSON file not found: {missing_json}')):
        check_known_items_cli.main([str(missing_json)])
